import os
from binance.client import Client
from datetime import datetime
import time
from openpyxl import Workbook
import telebot

# Ключи для Бинанс
api_key = ('Ваш_api_key')
api_secret = ('Ваш_api_secret')
client = Client(api_key, api_secret)

# Создаем таблицу для сбора данных
wb = Workbook()
ws = wb.active
# Расписываем названия колонок для данных
ws['A1'] = 'Average'
ws['B1'] = 'Futures'
ws['C1'] = 'Percentage'
ws['D1'] = 'Time'

# Задаём торговую пару:
ASSET = 'BTCUSDT'


def price(symbol):
    try:
        price = client.get_avg_price(symbol=symbol, requests_params={'timeout': 2})['price']
        return float(price)
    except Exception as e:
        print(e)


def priceF(symbol):
    try:
        priceF = client.futures_symbol_ticker(symbol=symbol, requests_params={'timeout': 2})['price']
        return float(priceF)
    except Exception as e:
        print(e)


bot = telebot.TeleBot('ваш_ключ_бота_от_BotFather')

# Ваш ID или список ID получателей
ID = 111111111


# Функция отправки сообщений
def message(text):
    bot.send_message(ID, text)


def message_signal():
    message(f'Сигнал по торговой паре: {ASSET}')


TIME = 1

# Процент выше которого мы начинаем получать сигналы
GROWTH_PERCENT = 0.053

while True:
    # Сравниваем цену фьючерсов со средней ценой
    FIRST_PRICE = price(ASSET)
    PRICEF = priceF(ASSET)
    PERCENT = ((PRICEF - FIRST_PRICE) / FIRST_PRICE) * 100

    if PERCENT >= GROWTH_PERCENT:
        # Нет необходимости получать результат в консоль, но я их обычно вывожу
        message_signal()
        print(ASSET)
        print("LOOK")
        print(PERCENT)
        times = datetime.now().strftime("%H:%M:%S")
        print(times)
        ws.append([FIRST_PRICE, PRICEF, PERCENT, times])
        wb.save("data.xlsx")

    time.sleep(TIME)

