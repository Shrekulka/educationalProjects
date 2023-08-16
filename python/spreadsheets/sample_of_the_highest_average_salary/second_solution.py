import openpyxl
import statistics

# Загрузка данных из файла
wb = openpyxl.load_workbook('salaries.xlsx')
sheet = wb.active

# Список для хранения медианных зарплат для каждого региона
median_salaries = []

# Проходим по каждой строке в таблице (строки 2-9) и вычисляем медианную зарплату для каждого региона
for row in sheet.iter_rows(min_row=2, max_row=9, values_only=True):
    region = row[0]
    salary_values = row[1:]  # Изменяем эту часть кода

    if region not in median_salaries:
        median_salaries.append((region, salary_values))  # Записываем название региона и зарплаты в список

# Находим регион с максимальной медианной зарплатой
max_median_region = max(median_salaries, key=lambda x: statistics.median(x[1]))[0]

# Список для хранения средних зарплат для каждой профессии
average_salaries = []

# Получаем названия профессий из первой строки (заголовки) и вычисляем среднюю зарплату для каждой профессии
professions = [cell.value for cell in sheet[1][1:]]
for idx, profession in enumerate(professions):
    salary_values = [row[idx + 1] for row in sheet.iter_rows(min_row=2, max_row=9, values_only=True)]  # Изменяем эту часть кода
    average_salaries.append((profession, statistics.mean(salary_values)))  # Записываем название профессии и среднюю зарплату в список

# Находим профессию с максимальной средней зарплатой
max_avg_profession = max(average_salaries, key=lambda x: x[1])[0]

# Выводим результаты
print(f"Регион с самой высокой медианной зарплатой: {max_median_region}")
print(f"Профессия с самой высокой средней зарплатой: {max_avg_profession}")
